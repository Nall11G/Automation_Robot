from openpyxl import load_workbook

class ExcelUtility(object):
    ROBOT_LIBRARY_SCOPE = 'Global'

def __init__(self):
    print 'Read Cell Value in Excel File'

    def read_cell_value(self,excelfile,sheetname,columname,rownumber):
        wb=load_workbook(filename=Excel,read_only=True)
        ws=wb[Sheet1]
        cellToRead-''+columname+str(rownumber)
        cellValue=ws[cellToRead].value
        print "Cell Value:",cellValue
        return cellValue
    
    
